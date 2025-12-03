#!/usr/bin/env python3
"""
Generate sample Excel file for Smart Expense Tracker
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta
except ImportError:
    print("Installing required package: openpyxl")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta

def create_sample_excel():
    """Create a sample Excel file with bank transactions"""

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Define headers
    headers = ["Date", "Description", "Amount", "Type", "Balance"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample transactions data
    # Format: (date_offset, description, amount, type)
    transactions = [
        # Income
        (0, "Salary for November", 50000, "CREDIT"),

        # Day 1 - Food & Groceries
        (1, "UPI/1234/Swiggy/Food Delivery", -250, "DEBIT"),
        (1, "POS 402912 DMART BLR IN", -1500, "DEBIT"),

        # Day 2 - Travel & Shopping
        (2, "UPI/5678/Uber/Morning Ride", -200, "DEBIT"),
        (2, "AMZ*Amazon Marketplace", -2500, "DEBIT"),

        # Day 3 - Food
        (3, "UPI/9012/Zomato/Lunch Order", -450, "DEBIT"),
        (3, "POS 234567 KFC RESTAURANT", -650, "DEBIT"),

        # Day 4 - Fuel & Entertainment
        (4, "ATM WDL HPCL PETROL PUMP BLR", -2000, "DEBIT"),
        (4, "UPI/3456/Netflix Subscription", -799, "DEBIT"),

        # Day 5 - Groceries & Travel
        (5, "POS 567890 RELIANCE FRESH", -1200, "DEBIT"),
        (5, "UPI/7890/OLA CABS/Evening", -150, "DEBIT"),

        # Day 6 - Shopping
        (6, "NEFT CR REFUND FROM FLIPKART", 500, "CREDIT"),
        (6, "UPI/2345/MYNTRA FASHION", -1999, "DEBIT"),

        # Day 7 - Groceries & Food
        (7, "UPI/4567/BIGBASKET/Vegetables", -2500, "DEBIT"),
        (7, "POS 789012 DOMINOS PIZZA", -800, "DEBIT"),

        # Day 8 - Bills
        (8, "Electricity Bill Payment BSES", -1800, "DEBIT"),
        (8, "Airtel Prepaid Recharge", -599, "DEBIT"),

        # Day 10 - Medical & Entertainment
        (10, "UPI/6789/APOLLO PHARMACY", -850, "DEBIT"),
        (10, "UPI/8901/BOOKMYSHOW/Movie", -500, "DEBIT"),

        # Day 12 - Rent
        (12, "House Rent Payment Nov", -15000, "DEBIT"),

        # Day 14 - Shopping & Food
        (14, "IMPS CR Cashback Credit", 100, "CREDIT"),
        (14, "UPI/1122/SWIGGY/Dinner", -350, "DEBIT"),

        # Day 15 - Groceries & Travel
        (15, "POS 345678 MORE SUPERMARKET", -1100, "DEBIT"),
        (15, "UPI/3344/UBER/Office", -180, "DEBIT"),

        # Day 17 - Fuel & Shopping
        (17, "ATM WDL BPCL PETROL PUMP", -1800, "DEBIT"),
        (17, "AMZ*Amazon Prime Membership", -1499, "DEBIT"),

        # Day 18 - Food & Bills
        (18, "UPI/5566/ZOMATO/Breakfast", -280, "DEBIT"),
        (18, "JIO Fiber Internet Bill", -999, "DEBIT"),

        # Day 20 - Medical & Entertainment
        (20, "POS 456789 MEDPLUS PHARMACY", -650, "DEBIT"),
        (20, "UPI/7788/HOTSTAR SUBSCRIPTION", -299, "DEBIT"),

        # Day 22 - Shopping & Food
        (22, "UPI/9900/FLIPKART/Electronics", -3500, "DEBIT"),
        (22, "POS 567890 PIZZA HUT", -750, "DEBIT"),

        # Day 24 - Groceries & Travel
        (24, "UPI/1212/BLINKIT/Groceries", -800, "DEBIT"),
        (24, "UPI/3434/OLA OUTSTATION", -1200, "DEBIT"),

        # Day 25 - Investment & Bills
        (25, "SIP MUTUAL FUND DEBIT", -5000, "DEBIT"),
        (25, "Water Bill Payment", -450, "DEBIT"),

        # Day 27 - Food & Entertainment
        (27, "UPI/5656/STARBUCKS COFFEE", -550, "DEBIT"),
        (27, "UPI/7878/YOUTUBE PREMIUM", -129, "DEBIT"),

        # Day 28 - Education & Shopping
        (28, "UDEMY COURSE PURCHASE", -799, "DEBIT"),
        (28, "UPI/9898/AJIO FASHION", -1599, "DEBIT"),

        # Day 30 - Final transactions
        (30, "IMPS CR Interest Credited", 250, "CREDIT"),
    ]

    # Starting date (November 1, 2024)
    start_date = datetime(2024, 11, 1)
    balance = 0

    # Add transactions
    for day_offset, description, amount, trans_type in transactions:
        trans_date = start_date + timedelta(days=day_offset)
        balance += amount

        row = [
            trans_date.strftime("%d-%m-%Y"),
            description,
            amount,
            trans_type,
            balance
        ]
        ws.append(row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Format amount and balance columns as currency
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Amount column (C)
        row[2].number_format = '#,##0.00'
        row[2].alignment = Alignment(horizontal="right")

        # Balance column (E)
        row[4].number_format = '#,##0.00'
        row[4].alignment = Alignment(horizontal="right")

        # Date alignment
        row[0].alignment = Alignment(horizontal="center")

        # Type alignment
        row[3].alignment = Alignment(horizontal="center")

    # Save file
    filename = "sample_bank_statement.xlsx"
    wb.save(filename)

    print(f"✅ Sample Excel file created: {filename}")
    print(f"\n📊 Summary:")
    print(f"   Total transactions: {len(transactions)}")
    print(f"   Date range: {start_date.strftime('%d-%m-%Y')} to {(start_date + timedelta(days=30)).strftime('%d-%m-%Y')}")
    print(f"   Final balance: ₹{balance:,.2f}")
    print(f"\n💡 Upload this file to test the expense tracker!")

    return filename

if __name__ == "__main__":
    create_sample_excel()

